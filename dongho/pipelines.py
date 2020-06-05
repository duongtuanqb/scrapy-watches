# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


class DonghoPipeline(object):
    def process_item(self, item, spider):
        return item


import os
import pandas
from openpyxl import load_workbook


class TgddExcelPipeline(object):
    def process_item(self, item, spider):
        file_path = 'data/{}.xlsx'.format(spider.name)
        file_is_exist = os.path.exists(file_path)
        columns = sorted(item.keys())
        df = pandas.DataFrame([item], columns=columns)
        writer = pandas.ExcelWriter(file_path, engine='openpyxl')

        if not file_is_exist:
            df.to_excel(
                writer,
                header=True,
                sheet_name=spider.name,
                index=False,
                columns=columns,
            )
            writer.save()
            writer.close()
        else:
            book = load_workbook(file_path)
            sheet = book.active
            max_row = sheet.max_row
            max_column = sheet.max_column

            header = [sheet.cell(row=1, column=i).value for i in range(1, max_column+1)]

            for col in columns:
                try:
                    sheet.cell(row=max_row + 1, column=header.index(col) + 1, value=str(item[col]))
                except:
                    sheet.cell(row=1, column=max_column + 1, value=str(col))
                    sheet.cell(row=max_row + 1, column=max_column + 1, value=str(item[col]))

            book.save(file_path)
            book.close()

        return item
